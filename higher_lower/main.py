import openpyxl
import random
from os import system, name


def compare_salaries(sal1,sal2):
    if sal1>sal2:
        return "V"
    return "N"


def first_display(sal, n):
    if n > 0:
        return f": {sal}Kč "
    return ""


def make_a_guess(sect):
    guess = ""
    while guess not in ['V', 'N']:
        guess = input(f"\nOdvětví {sect.upper()} má výšší či nížší mzdu? V/N: ").upper()
    return guess


header = ("\nPojďme si zahrát „Výš nebo níž“!"
      "\nUhodněte, ve kterém odvětví v roce 2023 byla vyšší průměrneá hrubá měsíční mzda.")

file = openpyxl.load_workbook('Průměrné hrubé měsíční mzdy podle odvětví - 2023.xlsx')
data = file["DATA"]

info = {}
sector = []
for i in range(1, data.max_row+1):
    key = data.cell(row=i, column=1).value
    sector.append(key)
    value = data.cell(row=i, column=2).value
    info[key] = value


game_on = True

while game_on:
    system('cls' if name == 'nt' else 'clear')
    print(header)
    random.shuffle(sector)
    score = 0
    i = 0

    while game_on and i < len(sector)-1:
        with open('best_score.txt', 'r') as best_score:
            best_score_now = best_score.readline()
            if best_score_now == "":
                best_score_now = 0

        salary = info[sector[i]]
        other_salary = info[sector[i+1]]
        answer = compare_salaries(other_salary, salary)
        displayed_salary = first_display(salary, i)

        print(f"\nPorovnejte:\n{sector[i].upper()}{displayed_salary}\na\n{sector[i+1].upper()}")
        higher_or_lower = make_a_guess(sector[i+1])

        system('cls' if name == 'nt' else 'clear')

        if answer == higher_or_lower:
            score += 1
            print(header)
            print(f"\nSpravně!")
            print(f"{sector[i].upper()}: {salary}Kč\n{sector[i + 1].upper()}: {other_salary}Kč")
            print(f"Skóre: {score}\nNejlepší skóre: {best_score_now}")
        else:
            game_on = False
            print("\nProhráli jste!")
            print(f"{sector[i].upper()}: {salary}Kč\n{sector[i + 1].upper()}: {other_salary}Kč")
        i += 1

        if score > int(best_score_now):
            with open('best_score.txt', 'w') as best_score:
                best_score.write(f"{score}")

    game_on = (input("\nJeště jsdnou? ano/ne: ").lower()=="ano")

system('cls' if name == 'nt' else 'clear')
print("\nDíky za hru!")