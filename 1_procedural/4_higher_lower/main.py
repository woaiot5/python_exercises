import openpyxl
import random
from os import system, name


def compare_salaries(sal1,sal2):
    if sal1>sal2:
        return "V"
    return "N"


def first_display(sal, n):
    if n > 0:
        return f": {sal}Kč "
    return ""


def make_a_guess(sect):
    guess = ""
    while guess not in ['V', 'N', 'EXIT']:
        guess = input(f"\nOdvětví {sect.upper()} má výšší či nížší mzdu? V/N: ").upper()
    system('cls' if name == 'nt' else 'clear')
    return guess


def best_score(read_or_write):
    global score, best_score_now
    if read_or_write == "read":
        with open('best_score.txt', 'r') as file_score:
            bs = file_score.readline()
            if bs == "":
                bs = 0
        return bs
    elif read_or_write == "write":
        if score > int(best_score_now):
            with open('best_score.txt', 'w') as file_score:
                file_score.write(f"{score}")
            return score
    return best_score_now


def result(true_false):
    global header, sector, i, salary, other_salary, score, best_score_now
    if true_false:
        print(header)
        print(f"\nSpravně!")
        print(f"{sector[i].upper()}: {salary}Kč\n{sector[i + 1].upper()}: {other_salary}Kč")
        print(f"Skóre: {score}\nNejlepší skóre: {best_score_now}")
    else:
        print("\nProhráli jste!")
        print(f"{sector[i].upper()}: {salary}Kč\n{sector[i + 1].upper()}: {other_salary}Kč")


header = ("\nPojďme si zahrát „Výš nebo níž“!"
      "\nUhodněte, ve kterém odvětví v roce 2023 byla vyšší průměrneá hrubá měsíční mzda."
          "\n(Pokud chcete hru ukončit, napište 'exit')")

file = openpyxl.load_workbook('Průměrné hrubé měsíční mzdy podle odvětví - 2023.xlsx')
data = file["DATA"]

info = {}
sector = []
for i in range(1, data.max_row+1):
    key = data.cell(row=i, column=1).value
    sector.append(key)
    value = data.cell(row=i, column=2).value
    info[key] = value


game_on = True

while game_on:
    system('cls' if name == 'nt' else 'clear')
    print(header)
    random.shuffle(sector)
    score = 0
    i = 0

    while game_on and i < len(sector)-1:
        best_score_now = best_score("read")
        salary = info[sector[i]]
        other_salary = info[sector[i+1]]
        answer = compare_salaries(other_salary, salary)
        displayed_salary = first_display(salary, i)

        print(f"\nPorovnejte:\n{sector[i].upper()}{displayed_salary}\na\n{sector[i+1].upper()}")
        higher_or_lower = make_a_guess(sector[i+1])

        if higher_or_lower == 'EXIT':
            game_on = False
        elif answer == higher_or_lower:
            score += 1
            best_score_now = best_score("write")
            result(True)
        else:
            game_on = False
            result(False)
        i += 1

    if score == len(sector)+1:
        print("\nUhadli jste vše spravně!")

    if higher_or_lower != 'EXIT':
        game_on = (input("\nJeště jsdnou? ano/ne: ").lower()=="ano")

system('cls' if name == 'nt' else 'clear')
print("\nDíky za hru!")